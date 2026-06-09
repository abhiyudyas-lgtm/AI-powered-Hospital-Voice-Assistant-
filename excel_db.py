from openpyxl import Workbook, load_workbook
import os
import threading

FILE = "appointments_main.xlsx"


def _write(name, disease, time):

    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Patient Name", "Disease", "Slot Time"])
        wb.save(FILE)

    wb = load_workbook(FILE)
    ws = wb.active
    ws.append([name, disease, time])
    wb.save(FILE)


def save_appointment(name, disease, time):
    # run in background – Twilio not blocked
    t = threading.Thread(target=_write, args=(name, disease, time))
    t.start()
